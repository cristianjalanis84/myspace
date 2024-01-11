# -*- coding: utf-8 -*-
"""
Created on Mon Jul 31 15:54:00 2023

@author: calanis
"""
bgColor = []
valuee = []
bgColor2 = []
fgColor = []
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\calanis\Desktop\Satellite Operations Handbook (SOH) - Part 2 - Flight Operational Procedures_ar1\fops\0841-3900-6PCOP-027-G Dump Configuration.xlsx',data_only=True)
wb2 = wb["Data Table"]
fs = wb.active
for ii in range(1,20):
    cell_color = fs.cell(column=ii, row=7)
    valuee.append(cell_color.value)
    bgColor.append(cell_color.fill.bgColor.index)
    fgColor.append(cell_color.fill.fgColor.index)
    
    
    
    
    

# import openpyxl
# from openpyxl import load_workbook
# excel_file = 'color_codes.xlsx' 
# wb = load_workbook(excel_file, data_only = True)
# sh = wb['Sheet1']
# color_in_hex = sh['A2'].fill.start_color.index # this gives you Hexadecimal value of the color
# print ('HEX =',color_in_hex) 
# print('RGB =', tuple(int(color_in_hex[i:i+2], 16) for i in (0, 2, 4))) # Color in RGB